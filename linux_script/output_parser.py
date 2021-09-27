from openpyxl import Workbook
import re


output_r = re.compile(r"episode (\d+) rewards = (\d+[.]\d+)")
if __name__ =="__main__":
    write_wb = Workbook()
    write_ws = write_wb.create_sheet('rl output')
    # write_ws = write_wb.active

    file_path = input("output file path: ")

    row = 3
    col = 3
    with open(file_path, 'r') as f:
        while True:
            line = f.readline()
            if not line: break
            else:
                m_obj = output_r.match(line)
                if m_obj:
                    epi = int(m_obj.group(1))
                    result_value = float(m_obj.group(2))
                    write_ws.cell(row, col, epi)
                    write_ws.cell(row, col + 1, result_value)
                    row += 1

        write_wb.save("output.xlsx")
